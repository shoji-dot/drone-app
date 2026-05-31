"""
国土交通省 無人航空機飛行記録（様式）のEXCEL出力
テンプレートファイルをベースに飛行データを書き込む
"""
import os
import io
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime, timedelta

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template_flight.xlsx")

# データ行の開始行（5行目〜19行目が飛行記録15件分）
DATA_START_ROW = 5
DATA_END_ROW = 19

def write_cell(ws, cell_addr, value):
    from openpyxl.utils import coordinate_to_tuple
    row, col = coordinate_to_tuple(cell_addr)
    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= row <= merged.max_row and
                merged.min_col <= col <= merged.max_col):
            ws.cell(merged.min_row, merged.min_col).value = value
            return
    ws[cell_addr].value = value

def minutes_to_hhmm(minutes: float) -> str:
    if not minutes:
        return "0:00"
    h = int(minutes // 60)
    m = int(minutes % 60)
    return f"{h}:{m:02d}"

def build_flight_excel(
    drone_name: str,
    drone_serial: str,
    drone_remote_id: str,
    flights: list,
    nr: str = ""
) -> bytes:
    """
    flights: list of dict with keys:
        date, pilot_name, pilot_license, flight_purpose,
        takeoff_location, landing_location,
        start_time, end_time, flight_minutes, total_flight_minutes,
        safety_notes,
        squawk_date, squawk_detail, action_date, action_detail, confirmer
    """
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # ── ヘッダー ───────────────────────────────────────────
    # B2: 登録記号（型式・製造番号）
    reg_id = drone_remote_id or f"{drone_name}　{drone_serial or ''}"
    write_cell(ws, "C2", reg_id)

    # NR番号 (R2セルに "(NR. XX)" 形式)
    if nr:
        write_cell(ws, "R2", f"\n（NR.　{nr}）")

    # ── 飛行データ行（5〜19行） ────────────────────────────
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for i, f in enumerate(flights[:15]):  # 最大15件
        row = DATA_START_ROW + i

        # B-C: 飛行年月日
        date_str = f.get("date", "")
        ws.cell(row=row, column=2).value = date_str
        ws.cell(row=row, column=2).alignment = center

        # D-F: 飛行させた者の氏名（+ 技能証明番号）
        pilot = f.get("pilot_name", "")
        license_no = f.get("pilot_license", "")
        pilot_cell = f"{pilot}\n{license_no}" if license_no else pilot
        ws.cell(row=row, column=4).value = pilot_cell
        ws.cell(row=row, column=4).alignment = center

        # G: 飛行概要
        ws.cell(row=row, column=7).value = f.get("flight_purpose", "")
        ws.cell(row=row, column=7).alignment = left

        # H: 離陸場所
        ws.cell(row=row, column=8).value = f.get("takeoff_location", "")
        ws.cell(row=row, column=8).alignment = center

        # I-K: 着陸場所
        ws.cell(row=row, column=9).value = f.get("landing_location", "")
        ws.cell(row=row, column=9).alignment = center

        # L: 離陸時刻
        ws.cell(row=row, column=12).value = f.get("takeoff_time", "")
        ws.cell(row=row, column=12).alignment = center

        # M-N: 着陸時刻
        ws.cell(row=row, column=13).value = f.get("landing_time", "")
        ws.cell(row=row, column=13).alignment = center

        # O: 飛行時間
        ws.cell(row=row, column=15).value = minutes_to_hhmm(f.get("flight_minutes", 0))
        ws.cell(row=row, column=15).alignment = center

        # P: 総飛行時間
        ws.cell(row=row, column=16).value = minutes_to_hhmm(f.get("total_flight_minutes", 0))
        ws.cell(row=row, column=16).alignment = center

        # Q-R: 飛行の安全に影響のあった事項
        ws.cell(row=row, column=17).value = f.get("safety_notes", "")
        ws.cell(row=row, column=17).alignment = left

    # ── 記事（不具合）セクション（21〜25行） ───────────────
    squawk_flights = [f for f in flights if f.get("squawk_detail")]
    for i, f in enumerate(squawk_flights[:5]):
        row = 21 + i
        ws.cell(row=row, column=3).value  = f.get("squawk_date", "")
        ws.cell(row=row, column=5).value  = f.get("squawk_detail", "")
        ws.cell(row=row, column=9).value  = f.get("action_date", "")
        ws.cell(row=row, column=11).value = f.get("action_detail", "")
        ws.cell(row=row, column=18).value = f.get("confirmer", "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
